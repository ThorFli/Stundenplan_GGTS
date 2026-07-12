from openpyxl import load_workbook
from collections import defaultdict
import re, os

os.chdir("/Users/thorstenflierl/Library/Mobile Documents/com~apple~CloudDocs/Projekte/Stundenplan")

def norm(x):
    return "" if x is None else str(x).strip()

wb = load_workbook('Stundenplan_Vorlage_Wuensche_Lehrer.xlsx', data_only=True)

ws3 = wb['03_Klassen']
h = [str(c.value or '').strip() for c in ws3[1]]
class_teachers = {}
for row in ws3.iter_rows(min_row=2, values_only=True):
    d = {h[i]: row[i] if i < len(row) else None for i in range(len(h))}
    k = norm(d.get('Klasse'))
    kl = norm(d.get('Klassenlehrer'))
    if k and kl:
        class_teachers[k] = kl

print('CLASS_TEACHERS_FROM_SHEET:', class_teachers)

wb2 = load_workbook('stundenplan_cpsat_ergebnis.xlsx', data_only=True)
rows = [r for r in wb2['Klassenplaene'].iter_rows(min_row=2, values_only=True) if r[6] == 'OK']

print()
print('--- Mo-1 / Fr-5 Check ---')
for klasse, kl in sorted(class_teachers.items()):
    for day, block in [('Montag', '1'), ('Freitag', '5')]:
        slot = [r for r in rows if r[0] == klasse and r[1] == day and r[2] == block]
        teachers = sorted({r[5] for r in slot})
        ok = kl in teachers
        print(f'  {"OK  " if ok else "MISS"} {klasse} {day} B{block}  expected={kl}  got={teachers}')

print()
print('--- DAZ 1.4 Lehrer ---')
daz = [r for r in rows if r[0] == '1.4' and r[4] == 'Deutsch als Zweitsprache']
by_day = defaultdict(list)
for r in daz:
    by_day[r[1]].append((r[2], r[5]))
for day in ['Montag', 'Dienstag', 'Mittwoch', 'Donnerstag', 'Freitag']:
    print(f'  {day}: {sorted(by_day[day])}')
all_daz = sorted({r[5] for r in daz})
print(f'  Alle DAZ-Lehrer: {all_daz}')
print(f'  Anzahl DAZ-Stunden: {len(daz)}')
