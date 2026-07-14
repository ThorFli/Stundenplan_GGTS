"""
CP-SAT Stundenplan-Solver v2
============================

Angepasst an die geaenderten Regeln aus `Stundenplan_Vorlage_Wuensche_Lehrer.xlsx`.
Kompatibel mit Python 3.13.

Installation lokal:
    python3.13 -m pip install -r requirements_cp_sat.txt

Ausfuehrung:
    python3.13 stundenplan_cpsat_solver_v2.py \
        --input Stundenplan_Vorlage_Wuensche_Lehrer.xlsx \
        --output stundenplan_cpsat_ergebnis.xlsx \
        --time-limit 300

Wichtig:
- OR-Tools muss lokal installiert sein.
- In dieser Chat-Umgebung ist OR-Tools nicht installiert; der Code wurde hier syntaktisch geprueft,
  kann aber erst lokal mit installiertem OR-Tools den Plan berechnen.
"""
from __future__ import annotations

import argparse
import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from ortools.sat.python import cp_model
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "OR-Tools ist nicht installiert. Bitte lokal ausfuehren:\n"
        "  python3.13 -m pip install -r requirements_cp_sat.txt\n"
        "oder:\n"
        "  python3.13 -m pip install ortools openpyxl\n"
    ) from exc

DAYS = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag"]
BLOCK_ORDER = {"1": 1, "2": 2, "3": 3, "4": 4, "4a": 4, "4b": 5, "5": 6, "Leseband": 7, "UF1": 8, "UF2": 9, "GF1": 10, "GF2": 11, "Feedback": 12}
NO_TEACHING_MARKERS = ["kein Einsatz im Unterricht", "keine Unterrichtszuteilung", "langzeiterkrankt", "EZ bis"]

SUBJECT_ALIASES = {
    "Mathe": "Mathematik",
    "Mathematik": "Mathematik",
    "DaZ": "Deutsch als Zweitsprache",
    "Daz": "Deutsch als Zweitsprache",
    "Deutsch als Zweitsprache": "Deutsch als Zweitsprache",
    "Deutsch": "Deutsch",
    "Sachunterricht": "Sachunterricht",
    "Musik": "Musik",
    "Sport": "Sport",
    "Schwimmen": "Schwimmen",
    "Kunst": "Bildende Kunst",
    "Bildende Kunst": "Bildende Kunst",
    "Französisch": "Französisch",
    "Religion": "Religion",
    "Evangelische Religion": "Religion",
    "evangelische Religion": "Religion",
    "Förder": "Förderunterricht",
    "Förderunterricht": "Förderunterricht",
}

HARD_SUBJECT_ALIASES = {
    "evangelische religion": "Religion",
    "religion": "Religion",
    "deutsch": "Deutsch",
    "mathematik": "Mathematik",
    "sachunterricht": "Sachunterricht",
    "sport": "Sport",
    "schwimmen": "Schwimmen",
    "förderunterricht": "Förderunterricht",
    "förderstunde": "Förderunterricht",
    "förder": "Förderunterricht",
}

DAY_ALIASES = {
    "montag": "Montag",
    "dienstag": "Dienstag",
    "mittwoch": "Mittwoch",
    "donnerstag": "Donnerstag",
    "freitag": "Freitag",
}

CORE_BLOCKS = ["1", "2", "3", "4", "5"]

@dataclass(frozen=True)
class Slot:
    day: str
    group: str
    block: str
    begin: str
    end: str
    typ: str

@dataclass
class Teacher:
    kuerzel: str
    name: str
    max_week: int | None
    role: str
    depot: str
    comment: str
    no_teaching: bool = False
    availability: dict[str, set[str]] = field(default_factory=dict)
    hard: str = ""
    soft: str = ""
    weight: int = 10
    start_first: str = "neutral"
    preferred_subjects: set[str] = field(default_factory=set)
    disliked_subjects: set[str] = field(default_factory=set)
    preferred_grades: set[str] = field(default_factory=set)
    preferred_classes: set[str] = field(default_factory=set)
    hard_blocked_slots: dict[str, set[str]] = field(default_factory=dict)
    hard_required_slot_subjects: list[tuple[str, str, str, str]] = field(default_factory=list)
    hard_required_subject_classes: list[tuple[str, str, int]] = field(default_factory=list)

@dataclass
class Need:
    klasse: str
    fach: str
    hours: int
    resource: str
    fixed_teacher: str = ""

@dataclass
class LessonTask:
    id: int
    klasse: str
    fach: str
    resource: str
    fixed_slots: list[tuple[str, str]] = field(default_factory=list)
    fixed_teacher: str = ""
    note: str = ""

@dataclass
class Assignment:
    klasse: str
    fach: str
    teacher: str
    day: str
    block: str
    begin: str
    end: str
    status: str
    note: str = ""


def norm(x: Any) -> str:
    return "" if x is None else str(x).strip()


def parse_int(x: Any) -> int | None:
    if x is None:
        return None
    if isinstance(x, int):
        return x
    if isinstance(x, float) and x.is_integer():
        return int(x)
    s = str(x).strip()
    if not s or s == "/":
        return None
    m = re.search(r"\d+", s)
    return int(m.group(0)) if m else None


def split_values(x: Any) -> list[str]:
    s = norm(x)
    if not s:
        return []
    return [p.strip() for p in re.split(r"[,;\n]+", s) if p.strip()]


def read_table(ws) -> list[dict[str, Any]]:
    headers = [norm(c.value) for c in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(norm(v) for v in values):
            continue
        rows.append({headers[i]: values[i] if i < len(values) else None for i in range(len(headers))})
    return rows


def grade_of_class(klasse: str) -> str:
    return klasse.split(".")[0]


def group_for_grade(grade: str) -> str:
    return "1+4" if grade in {"1", "4"} else "2+3"


def block_rank(block: str) -> int:
    return BLOCK_ORDER.get(block, 99)


def normal_subjects(text: Any) -> set[str]:
    s = norm(text)
    found = set()
    if not s:
        return found
    for key, val in SUBJECT_ALIASES.items():
        if key.lower() in s.lower():
            found.add(val)
    for p in split_values(s):
        found.add(SUBJECT_ALIASES.get(p, p))
    return found


def parse_grades(text: Any) -> set[str]:
    s = norm(text)
    if not s:
        return set()
    grades = set(re.findall(r"(?<!\d)[1-4](?!\.\d)", s))
    grades.update(m.group(1) for m in re.finditer(r"\b([1-4])\.\d\b", s))
    return grades


def parse_classes(text: Any) -> set[str]:
    return set(re.findall(r"\b[1-4]\.\d\b", norm(text)))


def _extract_blocks(fragment: str) -> set[str]:
    text = norm(fragment)
    blocks = set()
    for m in re.finditer(r"\b([1-5])\.?\s*bis\s*([1-5])\.?", text, flags=re.IGNORECASE):
        a = int(m.group(1))
        b = int(m.group(2))
        for i in range(min(a, b), max(a, b) + 1):
            blocks.add(str(i))
    for token in ["GF1", "GF2", "UF1", "UF2", "Leseband"]:
        if re.search(rf"\b{re.escape(token)}\b", text, flags=re.IGNORECASE):
            blocks.add(token)
    for m in re.finditer(r"\b([1-5])\.?\b", text):
        blocks.add(m.group(1))
    return blocks


def _norm_hard_subject(s: str) -> str:
    key = norm(s).lower()
    return HARD_SUBJECT_ALIASES.get(key, SUBJECT_ALIASES.get(norm(s), norm(s)))


def parse_hard_wishes(text: Any) -> tuple[dict[str, set[str]], list[tuple[str, str, str, str]], list[tuple[str, str, int]]]:
    raw = norm(text)
    if not raw:
        return defaultdict(set), [], []

    low = raw.lower()
    blocked = defaultdict(set)
    required_slot_subjects: set[tuple[str, str, str, str]] = set()
    required_subject_classes: set[tuple[str, str]] = set()

    # Beispiel: "kein langer Tag am Mittwoch"
    for day_key, day in DAY_ALIASES.items():
        if re.search(rf"kein\s+langer\s+tag\s+am\s+{day_key}", low):
            blocked[day].update({"UF1", "UF2", "GF1", "GF2"})

    # Beispiel: "Mittwochs 1. bis 4. Stunde KOOP"
    for day_key, day in DAY_ALIASES.items():
        for m in re.finditer(rf"{day_key}\w*\s+([^,;]+?)\s+koop", low):
            blocked[day].update(_extract_blocks(m.group(1)))

    # Beispiel: "Dienstag 4. und 5. Stunde evangelische Religion"
    subject_keys = sorted(HARD_SUBJECT_ALIASES.keys(), key=len, reverse=True)
    for day_key, day in DAY_ALIASES.items():
        for subj_key in subject_keys:
            pattern = (
                rf"{day_key}\w*\s+([^,;]+?)\s+{re.escape(subj_key)}"
                rf"(?:\s+(?:in\s+)?klasse\s*([1-4]\.\d)|\s+in\s+([1-4]\.\d))?"
            )
            for m in re.finditer(pattern, low):
                blocks = _extract_blocks(m.group(1))
                klasse = m.group(2) or m.group(3) or ""
                subject = _norm_hard_subject(subj_key)
                for block in blocks:
                    required_slot_subjects.add((day, block, subject, klasse))

    # Beispiel: "Deutsch Klasse 4.1" / "Sachunterricht in Klasse 4.3"
    for subj_key in subject_keys:
        subject = _norm_hard_subject(subj_key)
        for m in re.finditer(rf"{re.escape(subj_key)}\s+(?:in\s+)?klasse\s*([1-4]\.\d)", low):
            required_subject_classes.add((subject, m.group(1)))
        for m in re.finditer(rf"{re.escape(subj_key)}\s+in\s+([1-4]\.\d)", low):
            required_subject_classes.add((subject, m.group(1)))

    # Beispiel: "je eine Förderstunde in 4.1 und 4.3"
    for m in re.finditer(r"je\s+eine\s+[^,;]*förder[^,;]*\s+in\s+([1-4]\.\d)\s+und\s+([1-4]\.\d)", low):
        required_subject_classes.add((_norm_hard_subject("förderstunde"), m.group(1)))
        required_subject_classes.add((_norm_hard_subject("förderstunde"), m.group(2)))

    # DAZ-spezifische Muster: "DaZ Gruppe" gilt als Klasse ohne feste Nummer (leerer String = beliebig)
    # Muster 1: "jeden Tag X.-Y. Stunde in DaZ Gruppe" oder "jeden Tag X. Stunde in DaZ"
    # Beispiel: "jeden Tag 1.-3. Stunde in DaZ Gruppe"
    daz_subject = "Deutsch als Zweitsprache"
    for m in re.finditer(
        r"jeden\s+tag\s+(\d)\.?\s*(?:-|bis|und)\s*(\d)\.?\s*(?:stunden?|std)?\s+in\s+da[zZ]",
        low,
    ):
        b_from = int(m.group(1))
        b_to = int(m.group(2))
        for day in DAYS:
            for b in range(min(b_from, b_to), max(b_from, b_to) + 1):
                required_slot_subjects.add((day, str(b), daz_subject, ""))
    # Einzelner Block pro Tag: "jeden Tag X. Stunde in DaZ"
    for m in re.finditer(
        r"jeden\s+tag\s+(\d)\.?\s*(?:stunden?|std)?\s+in\s+da[zZ]",
        low,
    ):
        b = m.group(1)
        for day in DAYS:
            required_slot_subjects.add((day, b, daz_subject, ""))

    # Muster 2: "Dienstag, Mittwoch, Donnerstag, Freitag, N. Std in DaZ Gruppe"
    # Findet Folgen von Tagesnamen gefolgt von Block-Nr und "in DaZ"
    day_alts = "|".join(DAY_ALIASES.keys())
    for m in re.finditer(
        rf"((?:(?:{day_alts})\w*,?\s*)+)(\d)\.?\s*(?:stunden?|std)\s+in\s+da[zZ]",
        low,
    ):
        days_str = m.group(1)
        block = m.group(2)
        for day_key, day in DAY_ALIASES.items():
            if re.search(r"\b" + day_key + r"\b", days_str):
                required_slot_subjects.add((day, block, daz_subject, ""))

    required_subject_classes_list = [
        (subject, klasse, 1)
        for (subject, klasse) in sorted(required_subject_classes)
    ]
    return blocked, sorted(required_slot_subjects), required_subject_classes_list


def active_rules(wb) -> set[str]:
    if "13_Regeln" not in wb.sheetnames:
        return set()
    rules = set()
    for row in read_table(wb["13_Regeln"]):
        if norm(row.get("Aktiv")).lower() == "ja":
            rules.add(norm(row.get("Regel")))
    return rules


def load_data(path: Path):
    wb = load_workbook(path, data_only=True)
    rules = active_rules(wb)

    slots = []
    for row in read_table(wb["02_Zeiten"]):
        if norm(row.get("Planbar")).lower() != "ja":
            continue
        typ = norm(row.get("Typ"))
        if typ in {"Hofpause", "Fruehstueckspause Saal", "Feedbackrunde"}:
            continue
        slots.append(Slot(
            norm(row.get("Tag")), norm(row.get("Klassenstufen_Gruppe")), norm(row.get("Block")),
            norm(row.get("Beginn")), norm(row.get("Ende")), typ
        ))

    classes = []
    class_teachers = {}
    for row in read_table(wb["03_Klassen"]):
        k = norm(row.get("Klasse"))
        if not k:
            continue
        classes.append(k)
        kl = norm(row.get("Klassenlehrer"))
        if kl:
            class_teachers[k] = kl

    needs = []
    for row in read_table(wb["05_Unterrichtsbedarf"]):
        klasse = norm(row.get("Klasse"))
        fach = norm(row.get("Fach"))
        hours = parse_int(row.get("Wochenstunden")) or 0
        if not klasse or not fach or hours <= 0:
            continue
        res = norm(row.get("Ressource"))
        if res in {"0", "None"}:
            res = ""
        ft = norm(row.get("Lehrer"))
        if ft in {"0", "None"}:
            ft = ""
        needs.append(Need(klasse, fach, hours, res, ft))

    teachers = {}
    for row in read_table(wb["06_Lehrer"]):
        k = norm(row.get("Kuerzel"))
        if not k:
            continue
        comment = norm(row.get("Kommentar"))
        max_week = parse_int(row.get("Max_Std_Woche"))
        no_teaching = max_week is None or any(m.lower() in comment.lower() for m in NO_TEACHING_MARKERS)
        teachers[k] = Teacher(k, norm(row.get("Name")), max_week, norm(row.get("Rolle")), norm(row.get("Deputatsstunden_Blocke")), comment, no_teaching)

    if "07_Lehrer_Verfuegbarkeit" in wb.sheetnames:
        for row in read_table(wb["07_Lehrer_Verfuegbarkeit"]):
            k = norm(row.get("Kuerzel"))
            if k not in teachers:
                continue
            for day in DAYS:
                raw = norm(row.get(day))
                teachers[k].availability[day] = set() if raw.lower() == "frei" or not raw else set(split_values(raw))
    for t in teachers.values():
        for day in DAYS:
            t.availability.setdefault(day, set())

    if "10_Wuensche" in wb.sheetnames:
        for row in read_table(wb["10_Wuensche"]):
            k = norm(row.get("Kuerzel"))
            if k not in teachers:
                continue
            t = teachers[k]
            t.hard = norm(row.get("Harte_Wuensche"))
            t.soft = norm(row.get("Weiche_Wuensche"))
            t.weight = parse_int(row.get("Gewichtung")) or 10
            t.start_first = norm(row.get("Beginn_1_Stunde_JN")).lower() or "neutral"
            t.preferred_subjects = normal_subjects(row.get("Bevorzugte_Faecher"))
            t.disliked_subjects = normal_subjects(row.get("Nicht_Bevorzugte_Faecher"))

            # WICHTIG: Bevorzugte_Klassenstufen nur aus eigener Spalte lesen (kein Mix mit Freitext)
            pref_grade_text = norm(row.get("Bevorzugte_Klassenstufen"))
            t.preferred_grades = parse_grades(pref_grade_text)

            combined = " ".join([t.hard, t.soft, pref_grade_text])
            t.preferred_classes = parse_classes(combined)
            blocked, required_slots, required_classes = parse_hard_wishes(t.hard)
            t.hard_blocked_slots = blocked
            t.hard_required_slot_subjects = required_slots
            t.hard_required_subject_classes = required_classes
            if "kein einsatz im unterricht" in combined.lower():
                t.no_teaching = True
            if t.role == "Klassenlehrer" and k not in class_teachers.values():
                # Infer likely own class from strongest class preference if class sheet is empty.
                cls = sorted(t.preferred_classes)
                if cls:
                    class_teachers.setdefault(cls[0], k)

    return slots, classes, needs, teachers, class_teachers, rules


def slot_valid(slot: Slot, klasse: str, fach: str, rules: set[str]) -> bool:
    if slot.group != group_for_grade(grade_of_class(klasse)):
        return False
    if fach == "Gebundene Freizeit":
        return slot.block in {"GF1", "GF2"}
    if fach == "Ungebundene Freizeit":
        return slot.block in {"UF1", "UF2"}
    if fach == "Deutsch als Zweitsprache":
        return slot.block in {"1", "2", "3", "4"}
    if fach == "Leseband":
        return slot.block == "Leseband"
    if "Deutsch_Mathe" in rules and fach in {"Deutsch", "Mathematik"} and slot.block == "5":
        return False
    return slot.block in {"1", "2", "3", "4", "5"}


def teacher_available(t: Teacher, slot: Slot) -> bool:
    if t.no_teaching:
        return False
    if slot.block not in t.availability.get(slot.day, set()):
        return False
    if slot.block in t.hard_blocked_slots.get(slot.day, set()):
        return False
    dep = t.depot.lower()
    # Textual depot entries are not fully normalized; block same day/block if both appear.
    if slot.day.lower() in dep and slot.block.lower() in dep:
        return False
    return True


def create_tasks(needs: list[Need], classes: list[str]) -> list[LessonTask]:
    fixed_count = defaultdict(int)
    tasks = []
    i = 0

    # DAZ fixed Mo-Fr 1-4.
    if "1.4" in classes:
        for day in DAYS:
            for block in ["1", "2", "3", "4"]:
                tasks.append(LessonTask(i, "1.4", "Deutsch als Zweitsprache", "", [(day, block)], note="fix DAZ"))
                fixed_count[("1.4", "Deutsch als Zweitsprache")] += 1
                i += 1

    # Current Excel rule: Tuesday 4+5 Religion for all, Sport for one first class placeholder = 1.1.
    for klasse in classes:
        if klasse == "1.4":
            continue
        fach = "Sport" if klasse == "1.1" else "Religion"
        res = "Sporthalle" if fach == "Sport" else ""
        for block in ["4", "5"]:
            tasks.append(LessonTask(i, klasse, fach, res, [("Dienstag", block)], note="fix Religion/Sport Di 4+5"))
            fixed_count[(klasse, fach)] += 1
            i += 1

    for n in needs:
        remaining = max(0, n.hours - fixed_count[(n.klasse, n.fach)])
        for _ in range(remaining):
            tasks.append(LessonTask(i, n.klasse, n.fach, n.resource, [], n.fixed_teacher))
            i += 1
    return tasks


def compute_core_overload_by_class(tasks: list[LessonTask], classes: list[str]) -> dict[str, int]:
    core_count = defaultdict(int)
    for t in tasks:
        if t.fach in {"Gebundene Freizeit", "Ungebundene Freizeit", "Leseband"}:
            continue
        core_count[t.klasse] += 1

    overload = {}
    for klasse in classes:
        grade = grade_of_class(klasse)
        if grade in {"3", "4"}:
            overload[klasse] = max(0, core_count.get(klasse, 0) - 25)
        else:
            overload[klasse] = 0
    return overload


def select_overflow_task_ids(tasks: list[LessonTask], overload_by_class: dict[str, int]) -> set[int]:
    selected: set[int] = set()
    for klasse, overload in overload_by_class.items():
        if overload <= 0:
            continue

        eligible = [
            t for t in tasks
            if t.klasse == klasse
            and t.fach not in {"Gebundene Freizeit", "Ungebundene Freizeit", "Leseband"}
            and not t.fixed_slots
        ]
        eligible.sort(key=lambda t: (t.fixed_teacher != "", t.id))
        if len(eligible) < overload:
            raise ValueError(
                f"Unloesbar: Ueberhang fuer {klasse} kann nicht aufloesbar markiert werden (zu wenig flexible Stunden)."
            )
        for t in eligible[:overload]:
            selected.add(t.id)
    return selected


def pref_penalty(t: Teacher, klasse: str, fach: str, slot: Slot) -> int:
    p = 0
    txt = f"{t.hard} {t.soft}".lower()
    grade = grade_of_class(klasse)
    if klasse.lower() in txt or klasse in t.preferred_classes:
        p -= 35
    if grade in t.preferred_grades:
        p -= 20
    if fach in t.preferred_subjects:
        p -= 25
    if fach in t.disliked_subjects:
        p += 70
    if t.start_first == "nein" and slot.block == "1":
        p += 25
    if "später" in txt and slot.block == "1":
        p += 15
    if "kurz" in txt and slot.block in {"GF1", "GF2", "UF1", "UF2"}:
        p += 20
    if fach in {"Deutsch", "Mathematik", "Französisch", "Deutsch als Zweitsprache"} and block_rank(slot.block) > 4:
        p += 12
    return p


def solve_cp_sat(input_path: Path, output_path: Path, time_limit: int, workers: int):
    slots, classes, needs, teachers, class_teachers, rules = load_data(input_path)
    tasks = create_tasks(needs, classes)
    overload_by_class = compute_core_overload_by_class(tasks, classes)
    overflow_task_ids = select_overflow_task_ids(tasks, overload_by_class)
    task_by_id = {t.id: t for t in tasks}
    for task_id in overflow_task_ids:
        t = task_by_id[task_id]
        t.note = (t.note + " | " if t.note else "") + "Ueberhangstunde: Klassenlehrer legt diese Stunde auf GF"
    usable_teachers = {k: t for k, t in teachers.items() if not t.no_teaching}

    model = cp_model.CpModel()
    candidates = defaultdict(list)  # task_id -> [(var, slot, teacher, penalty)]
    class_slot_vars = defaultdict(list)
    teacher_slot_vars = defaultdict(list)
    resource_slot_vars = defaultdict(list)
    teacher_load = defaultdict(list)
    teacher_block_vars = defaultdict(list)
    class_subject_teacher_vars = defaultdict(dict)  # (klasse, fach) -> teacher -> y

    # NEU: echte 2-stufige Zielfunktion
    primary_objectives = []   # MUSS + UNASSIGNED
    secondary_objectives = [] # Soft-Ziele

    class_slot_entries = defaultdict(list)
    teacher_slot_entries = defaultdict(list)
    hard_wish_fallback_notes = []
    hard_wish_violations = []
    overflow_unassigned_vars = defaultdict(list)

    MUST_PREF_GRADE_WEIGHT = 300       # Verstoß Bevorzugte_Klassenstufen
    MUST_DISLIKED_SUBJECT_WEIGHT = 500   # Verstoß Nicht_Bevorzugte_Faecher
    PRIMARY_UNASSIGNED_WEIGHT = 1000
    SECONDARY_UNASSIGNED_WEIGHT = 20000
    HARD_WISH_PRIMARY_WEIGHT = 2000
    HARD_WISH_SECONDARY_WEIGHT = 50000
    NO_FREE_PRIMARY_WEIGHT = 1200
    NO_FREE_SECONDARY_WEIGHT = 25000
    MIN_LOAD_RATIO = 0.75              # Mindestauslastung 75% des Solls
    MIN_LOAD_PRIMARY_WEIGHT = 800
    MIN_LOAD_SECONDARY_WEIGHT = 15000

    # Nur diese Faecher duerfen notfalls leer bleiben:
    optional_unassigned_subjects = {"Ungebundene Freizeit", "Gebundene Freizeit"}

    for task in tasks:
        possible_slots = [s for s in slots if (not task.fixed_slots or (s.day, s.block) in task.fixed_slots) and slot_valid(s, task.klasse, task.fach, rules)]
        for slot in possible_slots:
            teacher_keys = list(usable_teachers)
            if task.fixed_teacher:
                teacher_keys = [task.fixed_teacher] if task.fixed_teacher in usable_teachers else []
            for k in teacher_keys:
                t = usable_teachers[k]
                if not teacher_available(t, slot):
                    continue
                if "Klassenlehrer_parallele_Foerderstunde" in rules and task.fach == "Förderunterricht":
                    kl = class_teachers.get(task.klasse, "")
                    if kl and k == kl:
                        continue
                    if not kl or kl not in usable_teachers:
                        continue
                    if not teacher_available(usable_teachers[kl], slot):
                        continue

                v = model.NewBoolVar(f"x_{task.id}_{task.klasse}_{task.fach}_{slot.day}_{slot.block}_{k}".replace(" ", "_"))
                pen = pref_penalty(t, task.klasse, task.fach, slot)
                candidates[task.id].append((v, slot, k, pen))
                class_slot_vars[(task.klasse, slot.day, slot.block)].append(v)
                teacher_slot_vars[(k, slot.day, slot.block)].append(v)
                teacher_block_vars[(k, slot.day, slot.block)].append(v)
                class_slot_entries[(task.klasse, slot.day, slot.block)].append((v, task, k))
                teacher_slot_entries[(k, slot.day, slot.block)].append((v, task))
                teacher_load[k].append(v)

                # Fach_Lehrer: ein Fach in einer Klasse soll (wochenweit) bei genau einem Lehrer liegen.
                # Foerderunt., DAZ sowie Freizeit-Bloecke (UF/GF) werden frei verteilt.
                if "Fach_Lehrer" in rules and task.fach not in {
                    "Förderunterricht", "Deutsch als Zweitsprache",
                    "Gebundene Freizeit", "Ungebundene Freizeit",
                }:
                    y_map = class_subject_teacher_vars[(task.klasse, task.fach)]
                    if k not in y_map:
                        y_map[k] = model.NewBoolVar(
                            f"y_{task.klasse}_{task.fach}_{k}".replace(" ", "_")
                        )
                    model.Add(v <= y_map[k])

                # Foerderunterricht mit Co-Lehrer: gleicher Task, zusaetzlich Klassenlehrer belegt.
                if "Klassenlehrer_parallele_Foerderstunde" in rules and task.fach == "Förderunterricht":
                    kl = class_teachers.get(task.klasse, "")
                    if kl and kl in usable_teachers and kl != k:
                        teacher_slot_vars[(kl, slot.day, slot.block)].append(v)
                        teacher_block_vars[(kl, slot.day, slot.block)].append(v)
                        teacher_slot_entries[(kl, slot.day, slot.block)].append((v, task))
                        teacher_load[kl].append(v)

                # MUSS-Logik: bevorzugte Klassenstufen (primär)
                g = grade_of_class(task.klasse)
                if t.preferred_grades and g not in t.preferred_grades:
                    primary_objectives.append(MUST_PREF_GRADE_WEIGHT * v)

                # MUSS-Logik: nicht bevorzugte Fächer vermeiden (primär)
                if task.fach in t.disliked_subjects:
                    primary_objectives.append(MUST_DISLIKED_SUBJECT_WEIGHT * v)

                if task.resource:
                    resource_slot_vars[(task.resource, slot.day, slot.block)].append(v)

                # Soft-Penalty nur sekundär
                if pen:
                    secondary_objectives.append(pen * v)

        # UNASSIGNED nur fuer UF/GF und explizite Ueberhangstunden erlauben.
        if task.fach in optional_unassigned_subjects or task.id in overflow_task_ids:
            u = model.NewBoolVar(f"unassigned_{task.id}_{task.klasse}_{task.fach}".replace(" ", "_"))
            candidates[task.id].append((u, None, "UNASSIGNED", 0))
            if task.id in overflow_task_ids:
                overflow_unassigned_vars[task.klasse].append(u)
            # Unbesetzte Stunden sollen nur im Notfall genutzt werden.
            primary_objectives.append(PRIMARY_UNASSIGNED_WEIGHT * u)
            secondary_objectives.append(SECONDARY_UNASSIGNED_WEIGHT * u)

    # Harte Vollbelegung (ausser UF/GF mit optionalem UNASSIGNED)
    for task in tasks:
        if not candidates[task.id]:
            raise ValueError(
                f"Unloesbar: Keine gueltigen Kandidaten fuer {task.klasse} / {task.fach} / {task.note}"
            )
        model.AddExactlyOne(v for v, _, _, _ in candidates[task.id])

    # Ueberhang 3./4. Klassenstufe: genau die Ueberhangmenge darf als extern auf GF gelegt gelten.
    for klasse, overload in overload_by_class.items():
        if overload <= 0:
            continue
        vars_k = overflow_unassigned_vars.get(klasse, [])
        if len(vars_k) < overload:
            raise ValueError(
                f"Unloesbar: Nicht genug Ueberhang-UNASSIGNED Variablen fuer {klasse} (benoetigt {overload})."
            )
        model.Add(sum(vars_k) == overload)

    # Harte_Wuensche als echte Nebenbedingungen (sofern parsebar).
    for k, t in usable_teachers.items():
        for day, block, subject, klasse in t.hard_required_slot_subjects:
            req_vars = []
            for task in tasks:
                if task.fach != subject:
                    continue
                if klasse and task.klasse != klasse:
                    continue
                for v, slot, tk, _p in candidates[task.id]:
                    if tk == k and slot and slot.day == day and slot.block == block:
                        req_vars.append(v)
            if not req_vars:
                # Fallback: Wunsch erkannt, aber kein passender Kandidat vorhanden.
                # GF/UF-Bloecke (Schwimm-Aufsicht etc.) sind kein regulaerer Unterricht -> nur Hinweis.
                msg = (
                    f"Hinweis {k}: {day} {block} {subject} {klasse} ist kein planbarer Unterrichtsslot "
                    f"(z.B. Schwimm-Aufsicht im GF-Block)."
                    if block in {"GF1", "GF2", "UF1", "UF2"}
                    else f"Harte_Wuensche Fallback {k}: {day} {block} {subject} {klasse} nicht direkt modellierbar."
                )
                hard_wish_fallback_notes.append(msg)
            else:
                # Hart: Lehrer MUSS diesen Slot fuer dieses Fach belegen (Harte_Wuensche).
                model.Add(sum(req_vars) >= 1)

        for subject, klasse, minimum in t.hard_required_subject_classes:
            req_vars = []
            for task in tasks:
                if task.fach == subject and task.klasse == klasse:
                    for v, slot, tk, _p in candidates[task.id]:
                        if tk == k and slot:
                            req_vars.append(v)
            if not req_vars:
                hard_wish_fallback_notes.append(
                    f"Harte_Wuensche Fallback {k}: {subject} in Klasse {klasse} nicht direkt modellierbar."
                )
            else:
                miss = model.NewBoolVar(
                    f"hardwish_miss_subject_{k}_{subject}_{klasse}_{minimum}".replace(" ", "_")
                )
                model.Add(sum(req_vars) + miss >= int(minimum))
                primary_objectives.append(HARD_WISH_PRIMARY_WEIGHT * miss)
                secondary_objectives.append(HARD_WISH_SECONDARY_WEIGHT * miss)
                hard_wish_violations.append((miss, f"Harte_Wuensche verletzt {k}: {subject} in Klasse {klasse}"))

    # Klassen-Slot-Kollisionen: immer max. 1 Eintrag je Klasse/Tag/Block.
    for vs in class_slot_vars.values():
        model.AddAtMostOne(vs)

    # Sport darf nicht parallel zu Foerderunterricht in derselben Klasse laufen.
    for key, entries in class_slot_entries.items():
        sport_vars = [v for v, t, _k in entries if t.fach == "Sport"]
        foerder_vars = [v for v, t, _k in entries if t.fach == "Förderunterricht"]
        if sport_vars and foerder_vars:
            model.Add(sum(sport_vars) + sum(foerder_vars) <= 1)

    # Keine Freistunden (weich, aber stark): in Block 1-5 moeglichst immer belegt,
    # ohne kuenstliche Faecher einzufuehren.
    for klasse in classes:
        for day in DAYS:
            for block in CORE_BLOCKS:
                slot_vars = [v for v, _t, _k in class_slot_entries.get((klasse, day, block), [])]
                if not slot_vars:
                    continue
                occupied = model.NewBoolVar(f"occupied_{klasse}_{day}_{block}".replace(" ", "_"))
                model.AddMaxEquality(occupied, slot_vars)
                free_slot = model.NewBoolVar(f"free_{klasse}_{day}_{block}".replace(" ", "_"))
                model.Add(occupied + free_slot == 1)
                primary_objectives.append(NO_FREE_PRIMARY_WEIGHT * free_slot)
                secondary_objectives.append(NO_FREE_SECONDARY_WEIGHT * free_slot)

    # Lehrer-Doppelbuchung: max. 1 pro Lehrer/Tag/Block.
    for vs in teacher_slot_vars.values():
        model.AddAtMostOne(vs)

    for vs in resource_slot_vars.values():
        model.AddAtMostOne(vs)

    # Klassenlehrer-Mitbelegung bei Foerderunterricht ist direkt in den Lehrer-Variablen modelliert.

    # Max_Std_Woche: Unterdeckung sekundär
    for k, t in usable_teachers.items():
        target = t.max_week
        if not target:
            continue
        vs = teacher_load.get(k, [])
        load_expr = sum(vs) if vs else 0
        model.Add(load_expr <= target)

        # Faire Mindestauslastung (weich): nach Moeglichkeit mind. 50% des Solls.
        min_target = max(1, int((target * MIN_LOAD_RATIO) + 0.9999))
        min_short = model.NewIntVar(0, min_target, f"min_shortfall_{k}")
        model.Add(min_short >= min_target - load_expr)
        primary_objectives.append(MIN_LOAD_PRIMARY_WEIGHT * min_short)
        secondary_objectives.append(MIN_LOAD_SECONDARY_WEIGHT * min_short)

        under = model.NewIntVar(0, target, f"underload_{k}")
        model.Add(under >= target - load_expr)
        secondary_objectives.append(1000 * under)

    if "Fach_Lehrer" in rules:
        for teacher_map in class_subject_teacher_vars.values():
            model.Add(sum(teacher_map.values()) == 1)

    # Sport als harte Doppelstunde, wenn Regel aktiv.
    if "Sport_moeglichst_Doppelstunde" in rules:
        allowed_patterns = [
            [0, 0, 0, 0, 0],
            [1, 1, 0, 0, 0],
            [0, 1, 1, 0, 0],
            [0, 0, 1, 1, 0],
            [0, 0, 0, 1, 1],
        ]
        for klasse in classes:
            if grade_of_class(klasse) == "1":
                continue
            sport_week = []
            for day in DAYS:
                day_bools = []
                for block in CORE_BLOCKS:
                    b = model.NewBoolVar(f"sport_{klasse}_{day}_{block}".replace(" ", "_"))
                    sport_vars = [
                        v for v, t, _k in class_slot_entries.get((klasse, day, block), [])
                        if t.fach == "Sport"
                    ]
                    if sport_vars:
                        model.AddMaxEquality(b, sport_vars)
                    else:
                        model.Add(b == 0)
                    day_bools.append(b)
                    sport_week.append(b)
                model.AddAllowedAssignments(day_bools, allowed_patterns)

            sport_required = sum(1 for t in tasks if t.klasse == klasse and t.fach == "Sport")
            model.Add(sum(sport_week) == sport_required)

    # Harte Regel: Wenn ein Fach an einem Tag mehrfach vorkommt, muessen die Stunden
    # aufeinanderfolgende Bloecke belegen (Doppelstunde). "Deutsch - Musik - Deutsch" ist verboten.
    # Gilt fuer alle Kernfaecher in Bloecken 1-5; nicht fuer Sport (eigene Doppelstunden-Regel),
    # DAZ (feste Slots), Foerderunterricht und Freizeit.
    _no_consecutive_check = {
        "Sport", "Förderunterricht", "Deutsch als Zweitsprache",
        "Gebundene Freizeit", "Ungebundene Freizeit", "Leseband",
    }
    for klasse in classes:
        faecher_klasse = {t.fach for t in tasks if t.klasse == klasse and t.fach not in _no_consecutive_check}
        for fach in faecher_klasse:
            for day in DAYS:
                # Bool-Variable pro Block: ist dieses Fach hier aktiv?
                block_active: dict[str, object] = {}
                for block in CORE_BLOCKS:
                    slot_vars_here = [
                        v for v, t, _k in class_slot_entries.get((klasse, day, block), [])
                        if t.fach == fach
                    ]
                    if slot_vars_here:
                        b = model.NewBoolVar(
                            f"consec_{klasse}_{fach}_{day}_{block}".replace(" ", "_")
                        )
                        model.AddMaxEquality(b, slot_vars_here)
                        block_active[block] = b

                if len(block_active) < 2:
                    continue  # hoechstens einmal an diesem Tag moeglich → kein Constraint noetig

                # Nicht-aufeinanderfolgende Blockpaare duerfen nicht beide belegt sein
                block_list = [b for b in CORE_BLOCKS if b in block_active]
                for ii, bi in enumerate(block_list):
                    for jj, bj in enumerate(block_list):
                        if jj <= ii + 1:
                            continue  # gleicher oder direkt folgender Block → erlaubt
                        # bi und bj sind nicht benachbart → duerfen nicht beide True sein
                        model.Add(block_active[bi] + block_active[bj] <= 1)

    # New rule: distribute subjects over weekdays as much as possible.
    if "Faecher_verteilt" in rules:
        for klasse in classes:
            subjects = {t.fach for t in tasks if t.klasse == klasse}
            for fach in subjects:
                for day in DAYS:
                    day_vars = []
                    for task in tasks:
                        if task.klasse == klasse and task.fach == fach:
                            for v, slot, _k, _p in candidates[task.id]:
                                if slot and slot.day == day:
                                    day_vars.append(v)
                    if len(day_vars) > 1:
                        count = sum(day_vars)
                        excess = model.NewIntVar(0, len(day_vars), f"excess_{klasse}_{fach}_{day}".replace(" ", "_"))
                        model.Add(excess >= count - 1)
                        secondary_objectives.append(120 * excess)

    # Harte Regel: Klassenlehrer muss Montag Block 1 und Freitag Block 5 in seiner Klasse sein.
    if "Klassenlehrer_Wochen_Start_Ende" in rules:
        for klasse, teacher in class_teachers.items():
            if teacher not in usable_teachers:
                continue
            for day, block in [("Montag", "1"), ("Freitag", "5")]:
                ok_vars = []
                for task in tasks:
                    if task.klasse != klasse:
                        continue
                    for v, slot, k, _p in candidates[task.id]:
                        if slot and slot.day == day and slot.block == block and k == teacher:
                            ok_vars.append(v)
                if ok_vars:
                    # Hart: Klassenlehrer MUSS an diesem Slot in seiner Klasse sein.
                    model.Add(sum(ok_vars) >= 1)
                else:
                    hard_wish_fallback_notes.append(
                        f"Klassenlehrer_Wochen_Start_Ende: {teacher} fuer {klasse} {day} B{block} nicht verfuegbar – Regel kann nicht eingehalten werden."
                    )

    # Nachmittagsmodell.
    late1545 = {"GF2"}
    late1445 = {"GF1", "GF2"}
    for k, t in usable_teachers.items():
        day1545 = []
        day1445 = []
        for day in DAYS:
            b1545 = model.NewBoolVar(f"late1545_{k}_{day}")
            b1445 = model.NewBoolVar(f"late1445_{k}_{day}")
            vs1545 = [v for (kk, dd, bb), vs in teacher_block_vars.items() if kk == k and dd == day and bb in late1545 for v in vs]
            vs1445 = [v for (kk, dd, bb), vs in teacher_block_vars.items() if kk == k and dd == day and bb in late1445 for v in vs]
            model.AddMaxEquality(b1545, vs1545) if vs1545 else model.Add(b1545 == 0)
            model.AddMaxEquality(b1445, vs1445) if vs1445 else model.Add(b1445 == 0)
            day1545.append(b1545)
            day1445.append(b1445)
        if t.role == "Klassenlehrer":
            model.Add(sum(day1545) <= 1)
            model.Add(sum(day1445) <= 2)
        elif t.role == "Fachlehrer":
            model.Add(sum(day1545) <= 2)

    # 2-Phasen-Optimierung:
    # Phase 1: primäre MUSS-Verstöße optimieren (nicht nur erste Lösung)
    # Phase 2: bei gegebener Primärschranke sekundäre Ziele optimieren
    primary_expr = sum(primary_objectives) if primary_objectives else 0
    secondary_expr = sum(secondary_objectives) if secondary_objectives else 0

    def _new_solver(tl=None, stop_after_first: bool = False):
        s = cp_model.CpSolver()
        s.parameters.max_time_in_seconds = float(tl if tl is not None else time_limit)
        s.parameters.num_search_workers = int(workers)
        s.parameters.log_search_progress = False
        if stop_after_first:
            s.parameters.stop_after_first_solution = True
        return s

    # Zeitbudget: 40 % fuer Phase 1 (Primaeroptimimerung), 60 % fuer Phase 2
    time_phase1 = max(30, int(time_limit * 0.4))
    time_phase2 = max(30, time_limit - time_phase1)

    best_primary = 0
    solver = None

    if primary_objectives:
        model.Minimize(primary_expr)
        # Phase 1: Primaeroziele wirklich optimieren
        solver = _new_solver(tl=time_phase1)
        status_primary = solver.Solve(model)

        if status_primary == cp_model.UNKNOWN:
            # Fallback: beliebige erste Loesung suchen, dann von dort aus optimieren
            solver = _new_solver(stop_after_first=True)
            status_primary = solver.Solve(model)

        if status_primary not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            status = status_primary
        else:
            best_primary = int(round(solver.ObjectiveValue()))
            for task in tasks:
                for v, _slot, _k, _p in candidates[task.id]:
                    model.AddHint(v, solver.Value(v))
            model.Add(primary_expr <= best_primary)
            model.Minimize(secondary_expr)
            solver = _new_solver(tl=time_phase2)
            status = solver.Solve(model)
    else:
        model.Minimize(secondary_expr)
        solver = _new_solver()
        status = solver.Solve(model)

    assignments = []
    warnings = []
    warnings.extend(hard_wish_fallback_notes)
    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        for miss, message in hard_wish_violations:
            if solver.Value(miss):
                warnings.append(message)
    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        for task in tasks:
            selected = None
            for v, slot, k, _p in candidates[task.id]:
                if solver.Value(v):
                    selected = (slot, k)
                    break
            if not selected:
                warnings.append(f"Keine Auswahl: {task.klasse} {task.fach}")
                continue
            slot, k = selected
            if slot is None:
                assignments.append(Assignment(task.klasse, task.fach, "UNASSIGNED", "", "", "", "", "UNASSIGNED", task.note))
                warnings.append(f"Nicht platziert: {task.klasse} {task.fach} {task.note}")
            else:
                assignments.append(Assignment(task.klasse, task.fach, k, slot.day, slot.block, slot.begin, slot.end, "OK", task.note))
                if "Klassenlehrer_parallele_Foerderstunde" in rules and task.fach == "Förderunterricht":
                    kl = class_teachers.get(task.klasse, "")
                    if kl and kl != k and kl in usable_teachers:
                        assignments.append(
                            Assignment(
                                task.klasse,
                                task.fach,
                                kl,
                                slot.day,
                                slot.block,
                                slot.begin,
                                slot.end,
                                "OK",
                                "Co-Lehrer Förderunterricht",
                            )
                        )

    else:
        warnings.append(
            "Keine gueltige Loesung gefunden (INFEASIBLE/UNKNOWN). "
            "Es wurden keine Unterrichtszuteilungen exportiert."
        )

    # Warnungen bei Unterdeckung je Lehrkraft
    load_by_teacher = defaultdict(int)
    for a in assignments:
        if a.status == "OK" and a.teacher in usable_teachers:
            load_by_teacher[a.teacher] += 1

    # MUSS-Verletzungen protokollieren
    pref_grade_violations = 0
    disliked_subject_violations = 0

    for a in assignments:
        if a.status != "OK" or a.teacher not in usable_teachers:
            continue
        t = usable_teachers[a.teacher]

        g = grade_of_class(a.klasse)
        if t.preferred_grades and g not in t.preferred_grades:
            pref_grade_violations += 1
            warnings.append(
                f"MUSS-Verstoss Bevorzugte_Klassenstufen: Lehrer {a.teacher} in {a.klasse} (Stufe {g}), "
                f"erlaubt: {', '.join(sorted(t.preferred_grades))}."
            )

        if a.fach in t.disliked_subjects:
            disliked_subject_violations += 1
            warnings.append(
                f"MUSS-Verstoss Nicht_Bevorzugte_Faecher: Lehrer {a.teacher} mit Fach {a.fach}."
            )

    total_underload = 0
    for k, t in usable_teachers.items():
        target = t.max_week
        if not target:
            continue
        assigned = load_by_teacher.get(k, 0)
        if assigned < target:
            deficit = target - assigned
            total_underload += deficit
            who = f"{k} ({t.name})" if t.name else k
            warnings.append(f"Unterdeckung Lehrer {who}: {assigned}/{target} Std (fehlen {deficit}).")

    stats = {
        "status": solver.StatusName(status),
        "objective": solver.ObjectiveValue() if status in (cp_model.OPTIMAL, cp_model.FEASIBLE) else "",
        "best_bound": solver.BestObjectiveBound() if status in (cp_model.OPTIMAL, cp_model.FEASIBLE) else "",
        "primary_score": best_primary,
        "wall_time": round(solver.WallTime(), 2),
        "tasks": len(tasks),
        "warnings": len(warnings),
        "teacher_underload_hours": total_underload,
        "preferred_grade_must_violations": pref_grade_violations,
        "disliked_subject_must_violations": disliked_subject_violations,
        "active_rules": ", ".join(sorted(rules)),
    }
    write_output(output_path, assignments, warnings, stats, class_teachers, rules)
    return stats


def write_output(
    path: Path,
    assignments: list[Assignment],
    warnings: list[str],
    stats: dict[str, Any],
    class_teachers: dict[str, str],
    rules: set[str],
):
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    bad_fill = PatternFill("solid", fgColor="F4CCCC")
    thin = Side(style="thin", color="D9D9D9")

    def header(row):
        for c in row:
            c.fill = header_fill
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = Border(bottom=thin)

    def auto(ws, width=None):
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = width or min(70, max(12, max(len(str(c.value or "")) for c in ws[letter]) + 2))
            for c in ws[letter]:
                c.alignment = Alignment(vertical="top", wrap_text=True)

    ws = wb.active
    ws.title = "Klassenplaene"
    ws.append(["Klasse", "Tag", "Block", "Zeit", "Fach", "Lehrer", "Status", "Hinweis"])
    header(ws[1])
    classplan_rows: list[Assignment] = []
    if "Klassenlehrer_parallele_Foerderstunde" in rules:
        foerder_grouped = defaultdict(list)
        passthrough = []
        for a in assignments:
            if a.status == "OK" and a.fach == "Förderunterricht":
                foerder_grouped[(a.klasse, a.day, a.block, a.begin, a.end)].append(a)
            else:
                passthrough.append(a)
        classplan_rows.extend(passthrough)
        for (_klasse, _day, _block, _begin, _end), rows in foerder_grouped.items():
            teachers = sorted({r.teacher for r in rows if r.teacher and r.teacher != "UNASSIGNED"})
            base = rows[0]
            classplan_rows.append(
                Assignment(
                    base.klasse,
                    base.fach,
                    " + ".join(teachers) if teachers else base.teacher,
                    base.day,
                    base.block,
                    base.begin,
                    base.end,
                    base.status,
                    "Co-Lehrer Förderunterricht" if len(teachers) > 1 else base.note,
                )
            )
    else:
        classplan_rows = list(assignments)

    for a in sorted(classplan_rows, key=lambda x: (x.klasse, DAYS.index(x.day) if x.day in DAYS else 99, block_rank(x.block))):
        ws.append([a.klasse, a.day, a.block, f"{a.begin}-{a.end}" if a.begin else "", a.fach, a.teacher, a.status, a.note])
        if a.status != "OK":
            for c in ws[ws.max_row]:
                c.fill = bad_fill
    auto(ws)

    # Required matrix layout: weekdays horizontal, hours/blocks vertical underneath.
    ws2 = wb.create_sheet("Klassenmatrix")
    blocks = ["1", "2", "3", "4", "5", "Leseband", "UF1", "UF2", "GF1", "GF2"]
    classes = sorted({a.klasse for a in assignments}, key=lambda k: (int(grade_of_class(k)) if "." in k else 99, k))
    row = 1
    for klasse in classes:
        ws2.cell(row, 1, klasse).font = Font(bold=True, size=14)
        row += 1
        ws2.append(["Block"] + DAYS)
        header(ws2[row])

        slot_assignments = defaultdict(list)
        for a in assignments:
            if a.klasse == klasse and a.status == "OK":
                slot_assignments[(a.day, a.block)].append(a)

        lookup = defaultdict(list)
        for (day, block), entries in slot_assignments.items():
            if "Klassenlehrer_parallele_Foerderstunde" in rules:
                foerder_entries = [e for e in entries if e.fach == "Förderunterricht"]
                sport_entries = [e for e in entries if e.fach == "Sport"]
                if foerder_entries and not sport_entries:
                    foerder = foerder_entries[0]
                    teachers = [foerder.teacher]
                    kl = next((e.teacher for e in entries if e.teacher == class_teachers.get(klasse, "") and e.teacher != foerder.teacher), "")
                    if kl:
                        teachers.append(kl)
                    hidden_subjects = sorted({e.fach for e in entries if e.fach != "Förderunterricht"})
                    hidden_txt = f"\nparallel: {', '.join(hidden_subjects)}" if hidden_subjects else ""
                    lookup[(day, block)].append(f"Förderunterricht\n{' + '.join(teachers)}{hidden_txt}")
                    continue

            for e in entries:
                lookup[(day, block)].append(f"{e.fach}\n{e.teacher}".strip())

        for block in blocks:
            ws2.append([block] + ["\n---\n".join(lookup.get((day, block), [])) for day in DAYS])

        row = ws2.max_row + 2
    auto(ws2, 22)

    ws3 = wb.create_sheet("Lehrerplaene")
    ws3.append(["Lehrer", "Tag", "Block", "Zeit", "Klasse", "Fach", "Hinweis"])
    header(ws3[1])
    for a in sorted([x for x in assignments if x.teacher and x.teacher != "UNASSIGNED"], key=lambda x: (x.teacher, DAYS.index(x.day), block_rank(x.block))):
        ws3.append([a.teacher, a.day, a.block, f"{a.begin}-{a.end}", a.klasse, a.fach, a.note])
    auto(ws3)

    ws4 = wb.create_sheet("Sporthalle")
    ws4.append(["Tag", "Block", "Zeit", "Klasse", "Fach", "Lehrer"])
    header(ws4[1])
    for a in sorted([x for x in assignments if x.fach in {"Sport", "Schwimmen"} and x.status == "OK"], key=lambda x: (DAYS.index(x.day), block_rank(x.block))):
        ws4.append([a.day, a.block, f"{a.begin}-{a.end}", a.klasse, a.fach, a.teacher])
    auto(ws4)

    ws5 = wb.create_sheet("Validierung")
    ws5.append(["Typ", "Meldung"])
    header(ws5[1])
    for k, v in stats.items():
        ws5.append(["Statistik", f"{k}: {v}"])
    for w in warnings:
        ws5.append(["Warnung", w])
        for c in ws5[ws5.max_row]:
            c.fill = warn_fill
    auto(ws5, 100)

    wb.save(path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="Stundenplan_Vorlage_Wuensche_Lehrer.xlsx")
    parser.add_argument("--output", default="stundenplan_cpsat_ergebnis.xlsx")
    parser.add_argument("--time-limit", type=int, default=300)
    parser.add_argument("--workers", type=int, default=8)
    args = parser.parse_args()
    stats = solve_cp_sat(Path(args.input), Path(args.output), args.time_limit, args.workers)
    print("CP-SAT fertig")
    for k, v in stats.items():
        print(f"{k}: {v}")

if __name__ == "__main__":
    main()
